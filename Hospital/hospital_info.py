
from openpyxl import Workbook , load_workbook
from openpyxl.utils import get_column_letter

from patients import get_random_patients, choice_patients, get_small_distance
from datetime import datetime ,time
from distance import dist_location 
from send_email import send_email
from msg_show import *

morning = '06:00:00'
evening = '18:00:00'

H_wb = load_workbook('Data/Data_set_Hospital.xlsx')
ws = H_wb.active

################################################################
def all_hospitals(ws):
    list_hospital_all = []
    list_hospital = []
    for row in range(2,22):
        list_hospital=[]
        for col in range(1,8):
            char= get_column_letter(col)    
            list_hospital.append(ws[char+str(row)].value)
        list_hospital_all.append(list(list_hospital)) 
    return list_hospital_all 


all_hospital = all_hospitals(ws)


def get_Public_hospital(all_hospital):
    list_hospital_req = []
    for hospital in all_hospital:
        if hospital[2] == 'Government':
            list_hospital_req.append(hospital)
    return list_hospital_req

###################################################################
def get_Private_hospital(all_hospital):
    list_hospital_req = []
    for hospital in all_hospital:
        if hospital[2] == 'Private':
            list_hospital_req.append(hospital)
    return list_hospital_req


##################################################################
def get_all_hospital(all_hospital):
    list_hospital_req = []
    for hospital in all_hospital:
        list_hospital_req.append(hospital)
    return list_hospital_req
#####################################################################
get_random_patient = get_random_patients(choice_patients)

print(get_random_patient)
# [0] = id
# [1] = lan
# [2] = la
# [3] = health
# [4] = heart
# [5] = name 
# [6] = time


# print("************************xXx***********************")


def check(get_random_patient):
    Triage = get_random_patient[3]
    diseases = get_random_patient[4]
    name = get_random_patient[5]
    time = get_random_patient[6]
    morning = '06:00:00'
    evening = '20:00:00'
    
    # Condition1: IF (time= evening AND triage (risk OR urgent)) THEN (hospital= (public AND send ambulance)).
    if time <= morning or time >=evening :
        print("Condition1")
        if Triage == 'risk' or Triage == 'Urgent':
            print(Triage)
            sugiest_hospital = get_Public_hospital(all_hospital)
            get_min_distance, min_dis = get_small_distance(sugiest_hospital, get_random_patient)
            hospital_Msg(name, Triage, diseases)
            patient_Msg_umbulance(name, Triage, diseases, get_min_distance[0][1])  # get_min_distance[0][1] = hospital name
            patientFamily_Msg(name, Triage, get_min_distance[0][1])
            send_email(get_min_distance[0][6], "There is a patient in the location " +
                       str(get_random_patient[1]) + " "+str(get_random_patient[2]))
            print("Condition1 - 1")
        # ******************************************
        elif Triage == 'Cold State' or Triage == 'Sick':
            print(Triage)
            sugiest_hospital = get_Public_hospital(all_hospital)
            get_min_distance, min_dis = get_small_distance(
                sugiest_hospital, get_random_patient)
            if Triage == 'Sick':
                patient_Msg_case2(name, Triage, diseases,
                                  get_min_distance[0][1])
                print("Condition2 - 1")
            elif Triage == 'Cold State':
                patient_Msg_case3(name, Triage, get_min_distance[0][1])
                send_email(get_min_distance[0][6], "There is a patient in the location " +
                           str(get_random_patient[1]) + " "+str(get_random_patient[2]))
                print("Condition2 - 2")

                # ******************************************
    # Condition2: IF(time=evening AND triage(cold case)) THEN(hospital=(public)).
    # elif time <= morning or time >= evening:
    #     print("Condition2")
    #     if Triage == 'Cold State' or Triage == 'Sick':
    #         print(Triage)
    #         sugiest_hospital = get_Public_hospital(all_hospital)
    #         get_min_distance, min_dis = get_small_distance(sugiest_hospital, get_random_patient)
    #         if Triage == 'Sick' :
    #             patient_Msg_case2(name, Triage, diseases, get_min_distance[0][1])
    #             print("Condition2 - 1")
    #         elif Triage == 'Cold State' :
    #             patient_Msg_case3(name, Triage, get_min_distance[0][1])
    #             send_email(get_min_distance[0][6], "There is a patient in the location " +
    #                        str(get_random_patient[1]) + " "+str(get_random_patient[2]))
    #             print("Condition2 - 2")
        
 # ******************************************
       # Condition3: IF (time= not evening AND triage (risk OR urgent)) THEN (hospital= (public OR private AND send ambulance)).
    elif time >= morning or time <= evening:
        print("Condition3")
        if Triage == 'risk' or Triage == 'Urgent':
            print(Triage)
            sugiest_hospital = get_all_hospital(all_hospital)
            get_min_distance, min_dis = get_small_distance(
                sugiest_hospital, get_random_patient)
            hospital_Msg(name, Triage, diseases)
            # get_min_distance[0][1] = hospital name
            patient_Msg_umbulance(name, Triage, diseases,
                                  get_min_distance[0][1])
            patientFamily_Msg(name, Triage, get_min_distance[0][1])
            send_email(get_min_distance[0][6], "There is a patient in the location " +
                       str(get_random_patient[1]) + " "+str(get_random_patient[2]))
            print("Condition3 - 1")
 # ******************************************
        elif Triage == 'Cold State' or Triage == 'Sick':
            print(Triage)
            sugiest_hospital = get_all_hospital(all_hospital)
            get_min_distance, min_dis = get_small_distance(
                sugiest_hospital, get_random_patient)
            if Triage == 'Sick':
                patient_Msg_case2(name, Triage, diseases,
                                  get_min_distance[0][1])
                print("Condition4 - 1")
            elif Triage == 'Cold State':
                patient_Msg_case3(name, Triage, get_min_distance[0][1])
                send_email(get_min_distance[0][6], "There is a patient in the location " +
                           str(get_random_patient[1]) + " "+str(get_random_patient[2]))
                print("Condition4 - 2")

         # ******************************************

    # Condition4: IF (time= not evening AND triage (cold case)) THEN (hospital= (public OR private)).
    # elif time >= morning or time <= evening:
    #     print("Condition4", Triage)
    #     if Triage == 'Cold State' or Triage == 'Sick':
    #         print(Triage)
    #         sugiest_hospital = get_all_hospital(all_hospital)
    #         get_min_distance, min_dis = get_small_distance(sugiest_hospital, get_random_patient)
    #         if Triage == 'Sick':
    #             patient_Msg_case2(name, Triage, diseases,get_min_distance[0][1])
    #             print("Condition4 - 1")
    #         elif Triage == 'Cold State':
    #             patient_Msg_case3(name, Triage, get_min_distance[0][1])
    #             send_email(get_min_distance[0][6], "There is a patient in the location " +
    #                        str(get_random_patient[1]) + " "+str(get_random_patient[2]))
    #             print("Condition4 - 2")
            
    else :
        print("time not found")

    # return sugiest_hospital


suggested_hospital = check(get_random_patient)


# def get_small_distance(sugiest_hospital):
#     min_dis=99999999999
#     list_small_dis=[]
#     for hospital_dis in sugiest_hospital:
#         if dist_location(hospital_dis[4], hospital_dis[5], get_random_patient[1], get_random_patient[2]) < min_dis:
#             list_small_dis = []
#             min_dis = dist_location(hospital_dis[4], hospital_dis[5], get_random_patient[1], get_random_patient[2])
#             list_small_dis.append(hospital_dis)
#     return list_small_dis , min_dis   





# get_min_distance , min_dis = get_small_distance(suggested_hospital)


# print(get_min_distance)

# print('small distance is ',min_dis)

# send_email(get_min_distance[0][6], "There is a patient in the location " +
#            str(get_random_patient[1]) + " "+str(get_random_patient[2]))









        
# 5.Condition1: IF (time= evening AND triage (risk OR urgent)) THEN (hospital= (public AND send ambulance)).
# 6.Condition2: IF (time= evening AND triage (cold case)) THEN (hospital = (public)).
# 7.Condition3: IF (time= not evening AND triage (risk OR urgent)) THEN (hospital= (public OR private AND send ambulance)).
# 8.Condition4: IF (time= not evening AND triage (cold case)) THEN (hospital= (public OR private)).




