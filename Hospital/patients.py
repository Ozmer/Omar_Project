from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import random
from datetime import datetime ,time
from distance import dist_location

def choice_patients():
    H_wb = load_workbook('Data/Data_set_P.xlsx')
    ws = H_wb.active
    list_patients_all = []
    list_patient = []
    for row in range(1,573):
        list_patient=[]
        for col in range(1,7):
            char = get_column_letter(col)
            list_patient.append(ws[char+str(row)].value)
        list_patients_all.append(list_patient)
    return random.choice(list_patients_all)


def get_random_patients(choice_patients):
    randon_patient = choice_patients()
    date_now = datetime.now()
    t = date_now.strftime("%H:%M:%S")
    randon_patient.append(t)
    return randon_patient


def get_small_distance(sugiest_hospital, get_random_patient):
    min_dis = 99999999999
    list_small_dis = []
    for hospital_dis in sugiest_hospital:
        if dist_location(hospital_dis[4], hospital_dis[5], get_random_patient[1], get_random_patient[2]) < min_dis:
            list_small_dis = []
            min_dis = dist_location(
                hospital_dis[4], hospital_dis[5], get_random_patient[1], get_random_patient[2])
            list_small_dis.append(hospital_dis)
    return list_small_dis, min_dis




